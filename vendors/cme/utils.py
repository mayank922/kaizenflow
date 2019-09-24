import logging
import os
import time

import numpy as np
import openpyxl
import pandas as pd
import requests
import xlrd
from bs4 import BeautifulSoup
from tqdm.autonotebook import tqdm

import helpers.dbg as dbg

_LOG = logging.getLogger(__name__)


class DownloadProductSlate:

    def __init__(self, download_url):
        """
        :param download_url: The url of the product slate
        """
        self.download_url = download_url

    def download_xls(self, dst_path):
        """
        Download product slate xls by url.

        :param dst_path: The path to save the xls.
        """
        dbg.dassert_eq(os.path.splitext(self.download_url)[1],
                       os.path.splitext(dst_path)[1],
                       "Extensions of the files do not match.")
        response = requests.get(self.download_url)
        if response.status_code == 200:
            with open(dst_path, 'wb') as f:
                f.write(response.content)
        else:
            raise ValueError(f'Request status code is {response.status_code}.')

    @staticmethod
    def open_xls_openpyxl(xls_path, first_row=1):
        """
        Open xls file into openpyxl (openpyxl does not support xls).
        You can choose from which row to save the file.

        :param xls_path: The path to the xls file
        :param first_row: From which row to open the file.
            If first_row==1, no rows are dropped (in openpyxl
            the numeration starts at 1)
        :return: openpyxl workbook
        """
        xlrd_book = xlrd.open_workbook(xls_path)
        xlrd_sheet = xlrd_book.sheet_by_index(0)
        openpyxl_wb = openpyxl.workbook.Workbook()
        openpyxl_sheet = openpyxl_wb.active
        for i, row in enumerate(range(first_row, xlrd_sheet.nrows)):
            for col in range(1, xlrd_sheet.ncols):
                openpyxl_sheet.cell(row=i + 1,
                                    column=col).value = xlrd_sheet.cell_value(
                    row, col)
        return openpyxl_wb

    @staticmethod
    def save_xls_to_xlsx(xls_path, first_row):
        """
        Open xls file, convert it to xlsx and save. You can choose from
        which row to save the file.

        :param xls_path: A path to the xls file
        :param first_row: From which row to open the file.
            If first_row==1, no rows are dropped (in openpyxl
            the numeration starts at 1).
        """
        openpyxl_wb = DownloadProductSlate.open_xls_openpyxl(xls_path=xls_path,
                                                             first_row=first_row)
        xlsx_path = os.path.splitext(xls_path)[0] + '.xlsx'
        openpyxl_wb.save(xlsx_path)

    def get_xlsx(self, dst_path, first_row=4):
        """
        A function that downloads xls, and saves the table to xlsx
        starting with the first_row row.

        :param dst_path: xls destination path
        :param first_row: From which row to save the file.
            If first_row==1, no rows are dropped (in openpyxl
            the numeration starts at 1).
        """
        self.download_xls(dst_path)
        self.save_xls_to_xlsx(xls_path=dst_path, first_row=first_row)


class ExtractHyperlinks:

    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path
        self.hyperlinks = []
        self.df_with_hyperlinks = pd.DataFrame()

    def extract(self, hyperlink_col_loc):
        """
        Get hyperlinks for the xlsx file.

        :param hyperlink_col_loc: From which column to extract hyperlinks.
            In openpyxl the numeration starts at 1.
        :return: A list of hyperlinks.
        """
        workbook = openpyxl.load_workbook(self.xlsx_path)
        ws = workbook[workbook.get_sheet_names()[0]]
        hyperlinks = []
        for row in ws.iter_rows(min_row=2, min_col=hyperlink_col_loc,
                                max_col=hyperlink_col_loc):
            hyperlinks.append(row[0].hyperlink.target)
        return hyperlinks

    def add_hyperlinks_to_df(self, hyperlink_col_name):
        """
        Read xlsx dataframe and add hyperlinks as a column.

        :param hyperlink_col_name: the name of the hyperlinks column
            in the output dataframe
        :return: pd.DataFrame with hyperlinks column
        """
        df = pd.read_excel(self.xlsx_path)
        df[hyperlink_col_name] = self.hyperlinks
        return df

    def get_df_with_hyperlinks(self, hyperlink_col_loc=5,
                               hyperlink_col_name='product_link'):
        """
        Extract hyperlinks from the dataframe and get a dataframe with
        a hyperlinks column.

        :param hyperlink_col_loc: Number of column from which to extract
            the hyperlinks. In openpyxl numeration starts at 1.
        :param hyperlink_col_name: The name of the hyperlinks column
            in the output dataframe
        :return: A dataframe with a hyperlinks column
        """
        self.hyperlinks = self.extract(hyperlink_col_loc=hyperlink_col_loc)
        self.df_with_hyperlinks = self.add_hyperlinks_to_df(
            hyperlink_col_name=hyperlink_col_name)

    def save_df_with_hyperlinks(self, dst_path, hyperlink_col_loc=5,
                                hyperlink_col_name='product_link'):
        """
        Extract hyperlinks from the dataframe and save a dataframe with
        a hyperlinks column.

        :param hyperlink_col_loc: Number of column from which to extract
            the hyperlinks. In openpyxl numeration starts at 1.
        :param hyperlink_col_name: The name of the hyperlinks column
            in the output dataframe
        """
        self.get_df_with_hyperlinks(hyperlink_col_loc=hyperlink_col_loc,
                                    hyperlink_col_name=hyperlink_col_name)
        self.df_with_hyperlinks.to_excel(dst_path)


class LoadHTML:

    def __init__(self):
        pass

    @staticmethod
    def extract_urls(soup_table):
        """
        Get hyperlinks from each cell of a beautiful soup table.

        :param soup_table: beautiful soup table
        :return: [[url]], the shape is (n_rows, n_cols). If there was no url
            in the cell, the corresponding element will be None
        """
        urls_in_rows = []
        for row in soup_table.find_all("tr"):
            urls_in_cols = []
            for td in row.find_all("td"):
                if td.find("a"):
                    url = td.a["href"]
                else:
                    url = None
                urls_in_cols.append(url)
            urls_in_rows.append(urls_in_cols)
        return urls_in_rows

    @staticmethod
    def _urls_list_to_df(urls_list):
        links_df = pd.DataFrame(urls_list).dropna(axis=1, how="all")
        links_df = links_df.add_prefix("link_")
        return links_df

    @staticmethod
    def extract_urls_df(soup_table):
        """
        Extract hyperlinks from a beautiful soup table.

        :param soup_table: beautiful soup table
        :return: pd.DataFrame with urls. If the column of the table had a
            URL, the output DataFrame will have this column with
            "link_" prefix. All elements of that column that do not have a
            link will be None.
        """
        return LoadHTML._urls_list_to_df(LoadHTML.extract_urls(soup_table))

    @staticmethod
    def soup_table_to_df(soup_table):
        """
        Read beautiful soup table to pandas DataFrame.

        :param soup_table: beautiful soup table
        :return: pd.DataFrame
        """
        return pd.read_html(str(soup_table).replace("colspan", ""))[0]

    @staticmethod
    def soup_table_to_df_with_links(soup_table):
        """
        Read beautiful soup table to pandas DataFrame. If a column contains
        hyperlinks, output dataframe will have this column with a prefix
        "link_"and extracted links. Elements that do not have a link
        will be None.

        :param soup_table: beautiful soup table
        :return: pd.DataFrame from that table with columns
            for extracted links.
        """
        df_without_links = LoadHTML.soup_table_to_df(soup_table)
        links_df = LoadHTML.extract_urls_df(soup_table)
        df_with_links = df_without_links.join(links_df)
        return df_with_links

    @staticmethod
    def load_html_to_df(html_url):
        """
        Download html by link, extract tables with hyperlinks from it,
        concatenate them into one dataframe.

        :param html_url: html link
        :return: pd.DataFrame of the tables extracted from the html.
            If a column has hyperlinks, they will be extracted into a
            column with a prefix "link_".
        """
        req_res = requests.get(html_url)
        if req_res.status_code == 200:
            soup = BeautifulSoup(req_res.content, "lxml")
            soup_tables = soup.find_all("table")
            dfs = [LoadHTML.soup_table_to_df_with_links(soup_table) for
                   soup_table in soup_tables]
            if len(dfs) > 0:
                concatenated_df = pd.concat(dfs)
            else:
                _LOG.info("No tables were extracted from %s", html_url)
                concatenated_df = None
        else:
            _LOG.warning(f"Request status code is {req_res.status_code}")
            concatenated_df = None
        return concatenated_df


class ContractSpecs:

    def __init__(self, product_slate):
        self.product_slate = product_slate
        self.name_link = self.product_slate.loc[:,
                         ['Product Name', 'product_link']].rename(
            columns={'Product Name': 'product_name'}).set_index('product_name')
        self.names_specs_dict = {}
        self.specs_df = pd.DataFrame()
        self.product_slate_with_specs = pd.DataFrame()

    def load_htmls(self):
        """
        Load htmls from links and construct a dictionary with them.

        :return: {product_name: html}
        """
        product_names_htmls = {}
        for name, link in tqdm(self.name_link.iterrows(),
                               total=len(self.name_link)):
            time.sleep(1)
            html = LoadHTML.load_html_to_df(link[0])
            if html is not None:
                html.set_index(0, inplace=True)
            product_names_htmls[name] = html
        return product_names_htmls

    def specs_dict_to_df(self):
        """
        Transform contract specs dictionary to a dataframe.

        :return: pd.DataFrame with product_name as index and contract
            spec html indices as columns
        """
        html_rows = [self._get_row(html,
                                   product_name) if html is not None else pd.DataFrame(
            index=[product_name]) for product_name, html in
                     self.names_specs_dict.items()]
        html_rows = [self._rename_duplicate_cols(html) for html in html_rows]
        specs_df = pd.concat(html_rows, sort=False)
        return specs_df

    def get_contract_specs(self):
        """
        Download contract specs and merge them to the product slate.

        :return: pd.DataFrame with product slate and contract specs
            for each product
        """
        self.names_specs_dict = self.load_htmls()
        self.specs_df = self.specs_dict_to_df()
        self.product_slate_with_specs = self.product_slate.merge(self.specs_df,
                                                                 left_on='Product Name',
                                                                 right_index=True)

    @staticmethod
    def _get_squash_cols(df):
        """
        Get names of all columns with int names that are above zero.

        :param df: pd.DataFrame
        :return: list of str column names
        """
        squash_cols = []
        for col_name in df.columns:
            if isinstance(col_name, int):
                if col_name > 0:
                    squash_cols.append(col_name)
        return squash_cols

    @staticmethod
    def _squash_cols(df, cols):
        if cols == [1]:
            squashed_series = df[1]
        else:
            squashed_series = df[cols].fillna('').apply(" ".join, axis=1)
        return squashed_series

    @staticmethod
    def _add_links_as_rows(df):
        link_df = df[['link_1']].dropna()
        link_df.columns = [1]
        series = pd.concat([df, link_df.rename('{} Link'.format)])[[1]]
        return series

    @staticmethod
    def _get_row(df, idx):
        df = df.copy()
        df[1] = ContractSpecs._squash_cols(df,
                                           ContractSpecs._get_squash_cols(df))
        df = df[[1, 'link_1']]
        tr_df = ContractSpecs._add_links_as_rows(df).T
        tr_df.index = [idx]
        return tr_df

    @staticmethod
    def _rename_duplicate_cols(df):
        df = df.copy()
        dupe_mask = df.columns.duplicated(keep='first')
        duped_col_names = [f"{col_name}_{i}" for i, col_name in
                           enumerate(df.columns[dupe_mask])]
        new_index = np.array(df.columns)
        new_index[dupe_mask] = duped_col_names
        df.columns = new_index
        return df
